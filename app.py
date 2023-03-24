from flask import Flask, request, render_template
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
import sqlite3
from openpyxl import load_workbook
import os, random
import string
import nltk
from nltk.tag import pos_tag

app = Flask(__name__)

# required for tokenization
nltk.download('punkt')
# required for parts of speech tagging
nltk.download('averaged_perceptron_tagger')
abbs = {"CC":"Coordinating Conjunction", 
        "CD":"Cardinal Digit", "DT":"Determiner", 
        "EX":"Existential like there is, there exists",
        "FW":"Foreign Word", 
        "IN":"Preposition or Subordinating Conjunction",
        "JJ":"Adjective", 
        "JJR":"Adjective, Comparative", 
        "JJS":"Adjective, Superlative",
        "LS":"List Marker", 
        "MD":"Modal", 
        "NN":"Noun, Singular", 
        "NNS":"Noun Plural",
        "NNP":"Proper Noun, Singular", 
        "NNPS":"Proper Noun, Plural", 
        "PDT":"Predeterminer",
        "POS":"Possessive Ending like parent’s", 
        "PRP":"Personal Pronoun like I, he, she",
        "PRP$":"Possessive Pronoun like my, his, hers",
        "PUNKT":"Punctuation",
        "RB":"Adverb like very, silently",
        "RBR":"Adverb, Comparative like better", 
        "RBS":"Adverb, Superlative like best",
        "RP":"Particle like give up", 
        "TO":"Preposition like go 'to' the store or Adverb like She pulled the door to behind her",
        "UH":"Interjection like er, mm", 
        "VB":"Verb, Base Form like take", 
        "VBD":"Verb, Past Tense like took", 
        "VBG":"Verb, Gerund or Present Participle like taking",
        "VBN":"Verb, Past Participle like taken", "VBP":"Verb, Sing Present like take",
        "VBZ":"Verb, 3rd person singular, present tense like takes",
        "WDT":"Determiner like which",
        "WP":"Pronoun like who, what",
        "WP$":"Possessive Pronoun lkie whose",
        "WRB":"Abverb like where, when"
}


thoughts = []
thoughts_counter = -1

# quotes list
def get_thoughts(filename):
    folder = os.path.join(os.getcwd(), 'static')
    f = open(folder+'/'+filename, 'r')
    global thoughts
    thoughts = f.readlines()
    #print(thoughts[0])
    f.close()

# quotes list
def get_quote(filename):
    folder = os.path.join(os.getcwd(), 'static')
    quotes = []
    f = open(folder+'/'+filename, 'r')
    quotes = f.readlines()
    f.close()
    quote = (random.choice(quotes)).split('|')
    return quote

def read_xl(xls):
    question_list = []
    folder = os.path.join(os.getcwd(), 'static')
    wb = load_workbook(folder + "/" + xls)
    sheet = wb["Sheet1"]
    row_count = sheet.max_row
    column_count = sheet.max_column
    for i in range(1, row_count):
        line = []
        for j in range(1, column_count + 1):
            data = sheet.cell(row=i+1, column=j).value
            line.append(str(data))
        question_list.append(line)
    return question_list

def load_mcq(filename, db):
    conn = openDB(db)
    lst = read_xl(filename)
    line_items = []
    for line in lst:
        items = line.split()
        line_str = [ i.join(' ') for i in items]
        line_items.append(line_str)

@app.route("/")
def main():
    quote = get_quote('quotes.txt')
    return render_template("index.html", quote=quote)

@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/people")
def memoirs():
    return render_template("people/index.html")

@app.route("/english")
def english():
    return render_template("english/index.html")

@app.route("/articles")
def article_posts():
    return render_template("articles/index.html")

@app.route("/blog")
def blog_posts():
    return render_template("blog/index.html")

@app.route("/questions")
def questions_posts():
    return render_template("questions/index.html")

@app.route("/helloworldpython")
def helloworld_posts():
    # return render_posts("templates/helloworldpython")
    return render_template("helloworldpython/index.html")

@app.route("/poems")
def poems_posts():
    return render_template("poems/index.html")

@app.route("/thoughts")
def thoughts_posts():
    return render_template('thoughts/index.html')

@app.route("/stories")
def stories_posts():
    return render_template("stories/index.html")

@app.route('/quiz', methods=['GET', 'POST'])
def quiz():
    if request.method == 'POST':
        f = request.files['file']
        f.save(os.path.join('static/quiz/', secure_filename(f.filename)))
    filenames = os.listdir('static/quiz')
    return render_template("quiz/index.html", filenames=filenames)

@app.route("/<category>/<title>", methods=['GET', 'POST'])
def show_pages(category, title):
    if category == 'quiz':
        return show_mcq(title)
    if category == 'thoughts':
        return show_thoughts(category, title)
    if category == 'facts':
        return show_fact(category, title)
    path = f'/{category}/{title}.html'
    return render_template(path)

def show_fact(category, title):
    try:
        conn = openDB('facts.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM factoid ORDER BY RANDOM() LIMIT 1;")
        row = cursor.fetchone(); 
        cursor.close()
    except sqlite3.Error as error:
        print("Failed to read single row from sqlite table", error)
    finally:
        if conn:
            conn.close()
            print("The SQLite connection is closed")
        path = f'/{category}/{title}.html'
        return render_template(path, row=row)

def show_thoughts(category, title):
    global thoughts
    if len(thoughts) == 0:
        get_thoughts("thoughts.txt")
    global thoughts_counter
    if thoughts_counter >= len(thoughts):
        thoughts_counter = -1
    thoughts_counter = thoughts_counter + 1
    thought = thoughts[thoughts_counter]
    page = f'/{category}/{title}.html'
    return render_template(page, thought=thought)

def show_mcq(mcq):
    lst = read_xl("quiz/"+mcq+".xlsx")
    return render_template('/quiz/mcq.html', lines=lst, title=mcq)

@app.route('/english/pos', methods=['GET', 'POST']) 
def pos():
    if request.method == 'POST':
        line = request.form.get('textField')
        items = get_tag_list(line)
        item_list = [{item[0]:item[1] for item in items}]            
        for item in item_list:
            for key in item:
                if key in string.punctuation:
                    item[key] = abbs['PUNKT']
                else:
                    item[key] = abbs[item[key]]

        return render_template('english/pos.html', items=item_list, line=line)
    else:
        return render_template('english/pos.html')
 
def get_tag_list(line):
    tokens = nltk.word_tokenize(line)

    # parts of speech tagging
    tagged = nltk.pos_tag(tokens)

    # print tagged tokens
    return tagged

##### Database ######
def openDB(db):
    conn = sqlite3.connect(db)
    return conn;

@app.route('/facts') 
def facts():
    return render_template('facts/index.html')

@app.route("/edit_fact")
def edit_fact():
    id = request.args.get("id", default=0, type=int)
    try:
        conn = openDB('facts.db')
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("selectb * from factoid where id = "+id )
        row = cur.fetchone();
        cur.close()
    except:
        msg = "error in fetch operation"
    finally:
        conn.close()
        return render_template("facts/edit_fact.html", row=row)

@app.route('/fact_list')
def fact_list():
    conn = openDB('facts.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("select * from factoid;")
    rows = cur.fetchall(); 
    conn.close()
    print(rows)
    return render_template("facts/fact_list.html", rows=rows)

@app.route("/add_fact", methods=['GET', 'POST'])
def addfactform(id):
    if request.method == 'POST':
        conn = openDB('facts.db')
        try:
            title = request.form['title']
            fact = request.form['fact']
            tag = request.form['tag']
            
            cur = conn.cursor()
            cur.execute("INSERT INTO factoid (title,fact,tag) VALUES (?,?,?)",(title,fact,tag) )
            conn.commit()
            msg = "Record successfully added"
        except:
            conn.rollback()
            msg = "error in insert operation"
      
        finally:
            conn.close()
            return fact_list()
    else:
        return render_template("facts/add_fact.html")


###
def render_posts(folder):
    posts_list = get_posts_list(folder)
    names = []
    posts = []
    for post in posts_list:
        name = post.split('/')[-1].split('.')[0]
        names.append(name)
        f = open(post,"r")
        lines = f.readlines()
        posts.append(' '.join(lines))
    return render_template("posts.html", posts=posts, names=names)

def get_posts_list(folder):
    folder = os.path.join(os.getcwd(), folder)
    # list to store files
    res = []
    # Iterate directory
    for path in os.listdir(folder):
        # check if current path is a file
        if os.path.isfile(os.path.join(folder, path)):
            res.append(os.path.join(folder, path))
    return res
###

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=8000)